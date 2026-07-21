"""
BSE Corporate Governance Report Downloader
===========================================
Scrapes https://www.bseindia.com/corporates/Corpgovernane.aspx using Playwright
and downloads the XBRL filing for every quarter available for each company.

The browser runs in visible (non-headless) mode to bypass Akamai bot detection.
Automation fingerprints are suppressed via launch flags and an init script.

Resume behaviour
----------------
- Each quarter is individually tracked in download_log.csv.
- On resume, only quarters with status "downloaded" are skipped; all previously
  failed quarters are retried regardless of whether the company folder exists.

Output
------
- governance_reports/<scrip_code>_<company_name>/   — one folder per company
  - XML files:  standard XBRL CG reports  (e.g. 532454_532454_CG.xml)
  - HTML files: iXBRL / ICGIG CG reports  (e.g. 532454_ICGIG_532454_CG.html)
  - Filenames use BSE's own suggested name prefixed with the scrip code.
- download_log.csv     — full audit trail (all statuses)
- failed_downloads.csv — subset of download_log.csv containing only failures
- downloader.log       — console-mirrored log file

Usage
-----
    # all companies
    python download_scripts/bse_cg_downloader.py --xlsx "data/raw/top_500_companies.xlsx"

    # first 100
    python download_scripts/bse_cg_downloader.py --xlsx "data/raw/top_500_companies.xlsx" --batch 1

    # next 100
    python download_scripts/bse_cg_downloader.py --xlsx "data/raw/top_500_companies.xlsx" --batch 2

Requirements
------------
    pip install playwright openpyxl tqdm
    playwright install chromium
"""

import re, csv, logging, argparse
from datetime import datetime
from pathlib import Path

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import load_workbook
from tqdm import tqdm



BASE_URL           = "https://www.bseindia.com/corporates/Corpgovernane.aspx"
_SCRIPT_DIR        = Path(__file__).resolve().parent          # .../DSL/download_scripts/
_DATA_DIR          = _SCRIPT_DIR.parent / "data"               # .../DSL/data/
_RAW_DIR           = _DATA_DIR / "raw"
_LOG_DIR           = _DATA_DIR / "logs"
OUTPUT_DIR         = _RAW_DIR / "governance_reports"
ZIP_PATH           = _RAW_DIR / "governance_reports.zip"
LOG_FILE           = _LOG_DIR / "cg_download_log.csv"
FAILED_FILE        = _LOG_DIR / "cg_failed_downloads.csv"
BATCH_SIZE_DEFAULT = 100

# ASP.NET server-control IDs on the CG page
ID_SEARCH   = "ContentPlaceHolder1_SmartSearch_smartSearch"
ID_HDN_CODE = "ContentPlaceHolder1_SmartSearch_hdnCode"   # hidden: scrip code for autocomplete
ID_HF_SCRIP = "ContentPlaceHolder1_hf_scripcode"          # hidden: scrip code for form submit
ID_PERIOD   = "ContentPlaceHolder1_ddlPeriod"
ID_SUBMIT   = "ContentPlaceHolder1_btnSubmit"
PERIOD_VAL  = "7"   # dropdown value for "Beyond last 1 year"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(), logging.FileHandler(_LOG_DIR / "cg_downloader.log")],
)
log = logging.getLogger(__name__)




def safe_dirname(name: str) -> str:
    """Strip characters that are illegal in directory names on Windows/macOS/Linux."""
    return re.sub(r'[<>:"/\\|?*]', '_', name).strip()


def load_scrip_codes(xlsx_path: str) -> list[tuple[int, str]]:
    """
    Read scrip codes and company names from the Excel file.

    Expects the active sheet to have scrip code in column B (index 1) and
    company name in column C (index 2), with a header row.

    Returns a list of (scrip_code, company_name) tuples.
    """
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    companies = [
        (int(row[1]), str(row[2]).strip())
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[1] and row[2]
    ]
    wb.close()
    log.info(f"Loaded {len(companies)} companies from {xlsx_path}")
    return companies


def load_done_set() -> set[str]:
    """
    Read download_log.csv and return a set of already-completed keys.

    Each key has the form "<scrip_code>|<quarter>" so individual quarters
    can be skipped without re-downloading the entire company.
    """
    done = set()
    if LOG_FILE.exists():
        with open(LOG_FILE, newline="") as f:
            for row in csv.DictReader(f):
                if row.get("status") == "downloaded":
                    done.add(f"{row['scrip_code']}|{row['quarter']}")
    return done


def append_log(row: dict) -> None:
    """
    Append one result row to download_log.csv.
    If the status is "failed", also append to failed_downloads.csv.

    Both files share the same schema:
        timestamp, scrip_code, company_name, quarter, filename, status, note
    """
    fields = ["timestamp", "scrip_code", "company_name",
              "quarter", "filename", "status", "note"]
    record = {"timestamp": datetime.now().isoformat(), **row}

    write_header = not LOG_FILE.exists()
    with open(LOG_FILE, "a", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        if write_header:
            w.writeheader()
        w.writerow(record)

    if row.get("status") == "failed":
        write_header = not FAILED_FILE.exists()
        with open(FAILED_FILE, "a", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            if write_header:
                w.writeheader()
            w.writerow(record)




def load_base_page(page) -> None:
    """
    Navigate to the CG landing page and block until the search form is ready.
    Logs the page title and a content snippet on timeout to help diagnose
    access-denied or redirect issues.
    """
    page.goto(BASE_URL, wait_until="domcontentloaded", timeout=30000)
    try:
        page.wait_for_selector(f"#{ID_SUBMIT}", timeout=30000)
    except PWTimeout:
        log.warning(f"Submit button not found. Page title: {page.title()!r}  URL: {page.url!r}")
        log.warning(f"Page snippet: {page.content()[:500]!r}")
        raise


def search_and_submit(page, scrip_code: int, company_name: str) -> None:
    """
    Submit a search for the given company without relying on the autocomplete widget.

    The autocomplete fires async requests that are hard to intercept reliably, so
    we bypass it by writing the scrip code directly into the two hidden ASP.NET
    fields that the form actually reads on submit. The period dropdown is set to
    "Beyond last 1 year" (value "7") on every call because navigating back to the
    base page resets all form controls.

    After clicking Submit, we wait for at least one table cell to appear rather
    than waiting for full network idle — this is significantly faster and robust
    to companies that have no filings (the selector simply times out silently).
    """
    page.fill(f"#{ID_SEARCH}", company_name)
    page.evaluate(f"""() => {{
        document.getElementById('{ID_HDN_CODE}').value = '{scrip_code}';
        document.getElementById('{ID_HF_SCRIP}').value = '{scrip_code}';
    }}""")
    page.select_option(f"#{ID_PERIOD}", value=PERIOD_VAL)
    page.click(f"#{ID_SUBMIT}")
    try:
        page.wait_for_selector("table tr td", timeout=15000)
    except PWTimeout:
        pass  # No results is fine — extract_table_rows will return []


def extract_table_rows(page) -> list[dict]:
    """
    Extract all filing rows from the results table in a single JavaScript call.

    Using a single JS evaluation avoids the per-element Playwright RPC round-trip
    cost that makes iterating rows from Python very slow.

    Returns a list of dicts with keys:
        scrip_code   — security code from column 1 (td index 0)
        company_name — security name from column 2 (td index 1)
        quarter      — quarter label from the anchor in column 4 (td index 3)
        xbrl_index   — this row's 0-based position among all <tr> elements with
                       <td> children; passed to download_xbrl to click the XBRL
                       link in column 6 (td index 5)
    """
    return page.evaluate("""() => {
        const rows = [];
        let idx = 0;
        for (const tr of document.querySelectorAll('table tr')) {
            const tds = tr.querySelectorAll('td');
            if (tds.length < 5) continue;
            const sc      = tds[0].innerText.trim();
            const qAnchor = tds[3].querySelector('a');
            const quarter = qAnchor ? qAnchor.innerText.trim() : '';
            if (sc && quarter) {
                rows.push({
                    scrip_code:   sc,
                    company_name: tds[1].innerText.trim(),
                    quarter:      quarter,
                    xbrl_index:   idx,
                });
            }
            idx++;
        }
        return rows;
    }""")


def _ext_from_bytes(data: bytes) -> str:
    """
    Detect the true file extension from magic bytes.

    BSE often serves XBRL files with an incorrect Content-Type of text/html,
    so we inspect the raw bytes instead of trusting the HTTP header.

    Detection order:
    - ZIP  : PK magic bytes (XBRL zip packages)
    - PDF  : %PDF magic bytes
    - HTML : UTF-8 BOM is stripped first, then content checked for <!DOCTYPE or
             <html — catches BOM-prefixed HTML pages (e.g. BSE session redirects)
    - XML  : any remaining content that starts with a tag, or has a bare UTF-8 BOM —
             all other markup is treated as XML since BSE XBRL files can begin with
             <xbrl, <submission, <Report, <?xml, etc.
    """
    if data[:4] == b"PK\x03\x04":
        return ".zip"
    if data[:4] == b"%PDF":
        return ".pdf"
    # Strip UTF-8 BOM if present, then normalise first 64 bytes for markup sniffing.
    # BOM must be stripped BEFORE the HTML check — a BOM-prefixed HTML page would
    # otherwise be misidentified as XML (the BOM check used to fire first).
    stripped = data.lstrip(b"\xef\xbb\xbf")
    head = stripped[:64].lstrip().lower()
    if head.startswith(b"<!doctype") or head.startswith(b"<html"):
        return ".html"
    if head.startswith(b"<") or data[:3] == b"\xef\xbb\xbf":
        return ".xml"
    return ".bin"


def download_xbrl(page, xbrl_index: int, comp_dir: Path,
                  scrip_code: int, q_safe: str):
    """
    Click the XBRL link for the given row and save the downloaded file.

    XBRL links on the BSE page use javascript: hrefs and must be clicked in
    the live browser. Two outcomes are handled:

    - Direct download (XML/ZIP): Playwright fires a "download" event when the
      server sets Content-Disposition: attachment. The file is saved using
      BSE's suggested filename, prefixed with the scrip code.
    - New tab (iXBRL HTML): Newer ICGIG filings open in a new browser tab.
      The page source is read directly from the open tab (new_tab.content())
      while the session is still active — re-fetching the URL via request.get()
      was unreliable because BSE checks the Referer header and returns a
      homepage redirect when it is absent. The filename is derived from the
      tab's URL, prefixed with the scrip code.

    Fallback filename CG_<quarter>_<scrip_code><ext> is used when no name
    can be determined from BSE's response.

    Returns:
        Path   — path of the saved file (success)
        None   — XBRL link exists but points to a dead/void href (not yet filed)
        False  — download attempted but could not be captured or saved
    """
    try:
        # The XBRL link is always the first anchor inside the 6th <td>
        data_rows = page.locator("table tr").filter(has=page.locator("td"))
        xbrl_link = data_rows.nth(xbrl_index).locator("td").nth(5).locator("a").first

        # A void/empty href means the XBRL file has not been filed yet
        href_val = xbrl_link.get_attribute("href") or ""
        if href_val.strip().lower() in ("javascript:void(0)", "javascript:;", "#", ""):
            log.info("  XBRL not yet available (dead link) — skipping")
            return None

        # Register both listeners BEFORE clicking so we catch whichever fires first
        downloads = []
        new_tabs  = []
        page.once("download", lambda dl: downloads.append(dl))
        page.context.once("page", lambda pg: new_tabs.append(pg))

        xbrl_link.click()

        # Poll in short intervals rather than a fixed sleep
        for _ in range(30):
            if downloads or new_tabs:
                break
            page.wait_for_timeout(300)

        # Case 1: browser triggered a file download.
        # Playwright only fires this event when the server sets Content-Disposition:
        # attachment — a session redirect never does this, so any content type is valid.
        if downloads:
            dl       = downloads[0]
            raw      = dl.path().read_bytes()
            ext      = _ext_from_bytes(raw)
            bse_name = dl.suggested_filename.strip()
            if bse_name:
                stem        = Path(bse_name).stem
                file_ext    = Path(bse_name).suffix or ext
                actual_dest = comp_dir / f"{scrip_code}_{stem}{file_ext}"
            else:
                actual_dest = comp_dir / f"CG_{q_safe}_{scrip_code}{ext}"
            actual_dest.parent.mkdir(parents=True, exist_ok=True)
            dl.save_as(actual_dest)
            log.info(f"  Saved as: {actual_dest.name}")
            return actual_dest

        # Case 2: link opened a new tab instead of triggering a download.
        # BSE iXBRL HTML filings open this way. Content is read directly from
        # the already-loaded tab rather than re-fetched via request.get() —
        # re-fetching loses the browser referrer/session context, which causes
        # BSE to return a homepage redirect or a malformed-header parse error.
        if new_tabs:
            new_tab = new_tabs[0]
            try:
                new_tab.wait_for_load_state("domcontentloaded", timeout=15000)
                new_url = new_tab.url

                if not new_url or new_url.startswith("about:"):
                    new_tab.close()
                else:
                    # Read rendered page source while the tab is still open and
                    # authenticated. For iXBRL HTML this is the exact file content.
                    html_str = new_tab.content()
                    new_tab.close()

                    body = html_str.encode("utf-8")
                    ext  = _ext_from_bytes(body)

                    # Detect session redirect: BSE homepage always has "Get Quote"
                    # in the first 4 KB; legitimate iXBRL files never do.
                    if ext == ".html" and b"Get Quote" in body[:4096]:
                        log.warning("  Response is BSE homepage (session redirect) — treating as failed")
                        return False

                    url_stem = Path(new_url.split("?")[0]).stem.strip()
                    actual_dest = (comp_dir / f"{scrip_code}_{url_stem}{ext}"
                                   if url_stem else
                                   comp_dir / f"CG_{q_safe}_{scrip_code}{ext}")
                    actual_dest.parent.mkdir(parents=True, exist_ok=True)
                    actual_dest.write_bytes(body)
                    log.info(f"  Saved as: {actual_dest.name}")
                    return actual_dest
            except Exception as e:
                log.warning(f"  New-tab handling error: {e}")
                try:
                    new_tab.close()
                except Exception:
                    pass

        log.warning(f"  Could not capture download for xbrl_index={xbrl_index}")
        return False

    except Exception as e:
        log.warning(f"  Download error: {e}")
        return False




def run(xlsx_path: str, batch: int = 0, batch_size: int = BATCH_SIZE_DEFAULT) -> None:
    """
    Main entry point.

    Loads the company list, slices the requested batch, then iterates through
    each company in a single persistent browser session (one page load per company
    rather than launching a new browser each time).

    Skip logic:
    - The whole run is skipped if governance_reports.zip already exists —
      that means finalize_dataset.py has already packaged a complete dataset.
      Delete the zip (or run a specific --batch) to re-scrape.
    - Otherwise, each quarter is skipped only if its scrip_code|quarter key is
      already present in download_log.csv with status "downloaded".
    - The company folder check is intentionally omitted so that failed quarters
      inside otherwise-partially-downloaded companies are retried.
    """
    if ZIP_PATH.exists():
        log.info(f"{ZIP_PATH} already exists — dataset already finalized, skipping scrape. "
                 "Delete it to re-run.")
        return

    companies = load_scrip_codes(xlsx_path)
    done_set  = load_done_set()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if batch > 0:
        start_idx = (batch - 1) * batch_size
        end_idx   = start_idx + batch_size
        todo      = companies[start_idx:end_idx]
        log.info(f"Batch {batch} — companies {start_idx + 1}–"
                 f"{min(end_idx, len(companies))} of {len(companies)}")
    else:
        todo = companies
        log.info(f"Processing all {len(todo)} companies")

    stats = {"downloaded": 0, "skipped": 0, "no_filing": 0, "failed": 0}

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=False,   # visible window required to bypass Akamai WAF
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
            ],
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/146.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 800},
        )
        # Suppress the navigator.webdriver flag that Akamai uses for bot detection
        context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            window.chrome = { runtime: {} };
        """)
        page = context.new_page()

        log.info("Loading BSE Corporate Governance page...")
        load_base_page(page)
        log.info("Page ready.")

        for scrip_code, company_name in tqdm(todo, desc="Companies"):
            comp_dir = OUTPUT_DIR / f"{scrip_code}_{safe_dirname(company_name)}"

            log.info(f"→ {scrip_code}  {company_name}")
            try:
                search_and_submit(page, scrip_code, company_name)
                rows = extract_table_rows(page)

                if not rows:
                    log.info("  No CG filings found")
                    stats["no_filing"] += 1
                    append_log({"scrip_code": scrip_code, "company_name": company_name,
                                "quarter": "", "filename": "", "status": "no_filing",
                                "note": "0 rows returned"})
                    load_base_page(page)
                    continue

                log.info(f"  Found {len(rows)} quarters")

                for row in rows:
                    key = f"{scrip_code}|{row['quarter']}"
                    if key in done_set:
                        stats["skipped"] += 1
                        continue

                    q_safe = re.sub(r'\s+', '_', row["quarter"])
                    result = download_xbrl(page, row["xbrl_index"],
                                          comp_dir, scrip_code, q_safe)

                    if result is None:
                        status   = "not_available"
                        filename = ""
                        stats["skipped"] += 1
                    elif result is False:
                        status   = "failed"
                        filename = ""
                        stats["failed"] += 1
                        log.warning(f"  ✗ FAILED: {row['quarter']}")
                    else:
                        status   = "downloaded"
                        filename = result.name
                        stats["downloaded"] += 1
                        done_set.add(key)
                        log.info(f"  ✓ {row['quarter']}  →  {filename}")

                    append_log({"scrip_code": scrip_code, "company_name": company_name,
                                "quarter": row["quarter"], "filename": filename,
                                "status": status, "note": ""})

                load_base_page(page)

            except Exception as e:
                log.error(f"  Unexpected error: {e}")
                stats["failed"] += 1
                try:
                    load_base_page(page)
                except Exception:
                    pass

        browser.close()

    log.info("=" * 60)
    log.info(
        f"DONE | Downloaded: {stats['downloaded']} | "
        f"Skipped: {stats['skipped']} | "
        f"No filing: {stats['no_filing']} | "
        f"Failed: {stats['failed']}"
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Download XBRL Corporate Governance filings from BSE India "
                    "for a list of companies in an Excel file.",
    )
    parser.add_argument("--xlsx", required = True,
                        help="Path to the Excel file containing scrip codes and company names.")
    parser.add_argument("--batch", type = int, default = 0,
                        help="1-indexed batch number to process. 0 (default) processes all companies.")
    parser.add_argument("--batch-size", type = int, default = BATCH_SIZE_DEFAULT,
                        help=f"Number of companies per batch (default: {BATCH_SIZE_DEFAULT}).")
    args = parser.parse_args()
    run(args.xlsx, args.batch, args.batch_size)